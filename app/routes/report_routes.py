from flask import Blueprint, jsonify, request, current_app, send_from_directory
from flask_jwt_extended import jwt_required, get_jwt, get_jwt_identity
from app.models import db
from app.models.generated_report_model import GeneratedReport
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

report_bp = Blueprint('report_bp', __name__, url_prefix='/api/reports')

@report_bp.route('/library', methods=['GET'])
@jwt_required()
def get_reports_library():
    # Meta-definition matching Procare library groups
    library = [
        {
            "id": "sign_in_out",
            "name": "Sign In-Out Students",
            "reports": [
                {"id": "student_attendance_summary", "name": "Attendance Summary", "description": "Provides total hours attended for all students."}
            ]
        },
        {
            "id": "clock_in_out",
            "name": "Clock In-Out Staff",
            "reports": [
                {"id": "staff_timesheet_summary", "name": "Staff Timesheet Summary", "description": "Total clock-in / clock-out hours worked."}
            ]
        },
        {
            "id": "billing",
            "name": "Billing",
            "reports": [
                {"id": "categorized_transaction_summary", "name": "Categorized Transaction Summary", "description": "Export all Invoices, Refunds, Credits, Subsidies, and Payments with ledger summaries."},
                {"id": "transactions", "name": "Transactions", "description": "Detailed history of charges and payments for a selected period."},
                {"id": "family_balance", "name": "Family Balance", "description": "Net outstanding balances across all active families."},
                {"id": "ledger_detailed_transaction", "name": "Ledger Detailed Transaction", "description": "Line-by-line detailed debit/credit changes per ledger account."},
                {"id": "default_payment_methods", "name": "Default Payment Methods", "description": "Lists default credit card or ACH tokens per family account."},
                {"id": "bank_deposit_slip", "name": "Bank Deposit Slip", "description": "Compile bank deposit logs for physical checks/cash deposits."},
                {"id": "billing_plan_tuition_report", "name": "Billing Plan Tuition Report", "description": "Active tuition plan allocations across the student registry."},
                {"id": "attendance_billing_plans", "name": "Attendance Billing Plans", "description": "Assigned hourly billing policies by student enrolment."},
                {"id": "aging_report", "name": "Aging Report", "description": "Accounts receivable aging summary (30/60/90+ days past due)."},
                {"id": "registration_payments", "name": "Registration Payments", "description": "Record of new student sign-up / registration fees."}
            ]
        }
    ]
    return jsonify(library), 200

@report_bp.route('/recent', methods=['GET'])
@jwt_required()
def get_recently_generated_reports():
    reports = GeneratedReport.query.order_by(GeneratedReport.created_at.desc()).limit(50).all()
    return jsonify([r.to_dict() for r in reports]), 200

@report_bp.route('/generate', methods=['POST'])
@jwt_required()
def generate_report():
    claims = get_jwt()
    role = claims.get('role')
    user_email = get_jwt_identity()
    
    created_by_id = None
    if role == 'superadmin':
        from app.models.super_admin_model import SuperAdmin
        admin = SuperAdmin.query.filter_by(email=user_email).first()
        if admin:
            created_by_id = admin.id
    elif role == 'staff':
        from app.models.staff_model import Staff
        staff = Staff.query.filter_by(email=user_email).first()
        if staff:
            created_by_id = staff.id

    data = request.get_json() or {}
    report_id = data.get('report_id')
    start_date_str = data.get('start_date')
    end_date_str = data.get('end_date')
    file_format = data.get('format', 'XLSX').upper()

    if not report_id or not start_date_str or not end_date_str:
        return jsonify({"error": "Missing required fields (report_id, start_date, end_date)."}), 400

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD."}), 400

    # Match Procare reports list
    report_names = {
        "categorized_transaction_summary": "Categorized Transaction Summary",
        "transactions": "Transactions Report",
        "family_balance": "Family Balance Report",
        "ledger_detailed_transaction": "Ledger Detailed Transaction Report",
        "default_payment_methods": "Default Payment Methods Report",
        "bank_deposit_slip": "Bank Deposit Slip",
        "billing_plan_tuition_report": "Billing Plan Tuition Report",
        "attendance_billing_plans": "Attendance Billing Plans",
        "aging_report": "Aging Report",
        "registration_payments": "Registration Payments"
    }
    
    report_name = report_names.get(report_id, "Custom Report")
    
    # Static files output
    reports_dir = os.path.join(current_app.root_path, 'static', 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ext = 'xlsx' if file_format == 'XLSX' else 'pdf'
    filename = f"{report_id}_{timestamp}.{ext}"
    output_path = os.path.join(reports_dir, filename)
    file_url = f"/api/reports/download/{filename}"

    try:
        if report_id == "categorized_transaction_summary":
            if file_format == 'XLSX':
                generate_categorized_transaction_summary(start_date, end_date, output_path)
            else:
                return jsonify({"error": "PDF format not implemented for this report yet. Please export as XLSX."}), 400
        else:
            # Mock generator for other library reports
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report_name[:30]
            ws["A1"] = f"{report_name} (Simulation)"
            ws["A2"] = f"Period: {start_date_str} to {end_date_str}"
            wb.save(output_path)
            
        # Log to database
        db_report = GeneratedReport(
            name=report_name,
            category="Billing" if report_id in report_names else "General",
            format=file_format,
            file_path=file_url,
            date_range=f"{start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}",
            created_by_id=created_by_id,
            created_by_role=role
        )
        db.session.add(db_report)
        db.session.commit()
        
        return jsonify(db_report.to_dict()), 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Failed to generate report: {e}")
        return jsonify({"error": f"Failed to generate report: {str(e)}"}), 500

@report_bp.route('/download/<path:filename>', methods=['GET'])
def download_report(filename):
    reports_dir = os.path.join(current_app.root_path, 'static', 'reports')
    return send_from_directory(reports_dir, filename, as_attachment=True)


# --- Helper Function for Categorized Transaction Summary ---
def generate_categorized_transaction_summary(start_date, end_date, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaction Summary"
    ws.views.sheetView[0].showGridLines = True

    # Styling helpers
    title_font = Font(name="Arial", size=14, bold=True, color="1E293B")
    section_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=9, bold=True, color="475569")
    body_font = Font(name="Arial", size=9, color="000000")
    bold_body_font = Font(name="Arial", size=9, bold=True, color="000000")
    
    section_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    summary_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Block
    ws.merge_cells("A1:E1")
    ws["A1"] = "Categorized Transaction Summary"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Date Range: {start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="64748B")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    
    current_row = 4

    def write_section_header(title):
        nonlocal current_row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    def write_table_headers(headers):
        nonlocal current_row
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # --- SECTION 1: INVOICES (DEBITS) ---
    write_section_header("1. INVOICES (CHARGES)")
    write_table_headers(["Date", "Invoice ID", "Student Name", "Description", "Amount"])
    
    from app.models.financial_model import Invoice
    invoices_in_range = Invoice.query.filter(Invoice.created_at >= start_date, Invoice.created_at <= end_date).all()
    invoice_total = 0.0
    
    for inv in invoices_in_range:
        student_name = f"{inv.account.student.first_name} {inv.account.student.last_name}" if inv.account and inv.account.student else "Unknown Student"
        for item in inv.items:
            amount = float(item.amount or 0)
            invoice_total += amount
            
            row_vals = [inv.created_at.strftime('%m/%d/%Y'), f"INV#{inv.id}", student_name, item.description, amount]
            for col_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = body_font
                cell.border = thin_border
                if col_idx == 5:
                    cell.number_format = "$#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
            current_row += 1
            
    # Invoice Total
    ws.cell(row=current_row, column=4, value="Invoices Total").font = bold_body_font
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=current_row, column=5, value=invoice_total)
    total_cell.font = bold_body_font
    total_cell.number_format = "$#,##0.00"
    total_cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    total_cell.alignment = Alignment(horizontal="right")
    current_row += 2

    # --- SECTION 2: REFUNDS (DEBITS) ---
    write_section_header("2. REFUNDS")
    write_table_headers(["Date", "Payment ID", "Student Name", "Description / Notes", "Amount"])
    
    from app.models.financial_model import Payment
    refunds_in_range = Payment.query.filter(Payment.method == 'Refund', Payment.transaction_date >= start_date, Payment.transaction_date <= end_date).all()
    refund_total = 0.0
    
    for ref in refunds_in_range:
        student_name = f"{ref.account.student.first_name} {ref.account.student.last_name}" if ref.account and ref.account.student else "Unknown Student"
        amount = float(ref.amount or 0)
        refund_total += amount
        
        row_vals = [ref.transaction_date.strftime('%m/%d/%Y'), f"REF#{ref.id}", student_name, ref.notes or "Refund", amount]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if col_idx == 5:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        current_row += 1
        
    # Refunds Total
    ws.cell(row=current_row, column=4, value="Refunds Total").font = bold_body_font
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=current_row, column=5, value=refund_total)
    total_cell.font = bold_body_font
    total_cell.number_format = "$#,##0.00"
    total_cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    total_cell.alignment = Alignment(horizontal="right")
    current_row += 2

    # --- SECTION 3: CREDITS (CREDITS) ---
    write_section_header("3. CREDITS (ADJUSTMENTS)")
    write_table_headers(["Date", "Credit ID", "Student Name", "Reason", "Amount"])
    
    from app.models.financial_model import Credit
    credits_in_range = Credit.query.filter(Credit.created_at >= start_date, Credit.created_at <= end_date).all()
    credit_total = 0.0
    
    for cred in credits_in_range:
        student_name = f"{cred.account.student.first_name} {cred.account.student.last_name}" if cred.account and cred.account.student else "Unknown Student"
        amount = float(cred.amount or 0)
        credit_total += amount
        
        row_vals = [cred.created_at.strftime('%m/%d/%Y'), f"CRE#{cred.id}", student_name, cred.reason, -amount]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if col_idx == 5:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        current_row += 1
        
    # Credits Total
    ws.cell(row=current_row, column=4, value="Credits Total").font = bold_body_font
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=current_row, column=5, value=-credit_total)
    total_cell.font = bold_body_font
    total_cell.number_format = "$#,##0.00"
    total_cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    total_cell.alignment = Alignment(horizontal="right")
    current_row += 2

    # --- SECTION 4: SUBSIDIES (CREDITS) ---
    write_section_header("4. SUBSIDY / SPONSOR DISTRIBUTIONS")
    write_table_headers(["Date", "Ref / Check #", "Student Name", "Notes", "Amount"])
    
    from app.models.subsidy_transaction_model import SubsidyPaymentDistribution, SubsidyTransaction
    subsidies_in_range = SubsidyPaymentDistribution.query.join(SubsidyTransaction).filter(
        SubsidyTransaction.transaction_date >= start_date.date(),
        SubsidyTransaction.transaction_date <= end_date.date()
    ).all()
    
    subsidy_total = 0.0
    for sub in subsidies_in_range:
        student_name = f"{sub.student_account.student.first_name} {sub.student_account.student.last_name}" if sub.student_account and sub.student_account.student else "Unknown Student"
        amount = float(sub.amount or 0)
        subsidy_total += amount
        
        row_vals = [
            sub.transaction.transaction_date.strftime('%m/%d/%Y'),
            sub.transaction.reference_number or "N/A",
            student_name,
            sub.transaction.notes or "Subsidy distribution",
            -amount
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if col_idx == 5:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        current_row += 1
        
    # Subsidy Total
    ws.cell(row=current_row, column=4, value="Subsidy Total").font = bold_body_font
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=current_row, column=5, value=-subsidy_total)
    total_cell.font = bold_body_font
    total_cell.number_format = "$#,##0.00"
    total_cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    total_cell.alignment = Alignment(horizontal="right")
    current_row += 2

    # --- SECTION 5: PAYMENTS (CREDITS) ---
    write_section_header("5. PAYMENTS")
    write_table_headers(["Date", "Payment ID", "Student Name", "Method & Notes", "Amount"])
    
    payments_in_range = Payment.query.filter(Payment.method != 'Refund', Payment.transaction_date >= start_date, Payment.transaction_date <= end_date).all()
    payment_total = 0.0
    
    for pay in payments_in_range:
        student_name = f"{pay.account.student.first_name} {pay.account.student.last_name}" if pay.account and pay.account.student else "Unknown Student"
        amount = float(pay.amount or 0)
        payment_total += amount
        
        desc = f"{pay.method}" + (f" - {pay.notes}" if pay.notes else "")
        row_vals = [pay.transaction_date.strftime('%m/%d/%Y'), f"PAY#{pay.id}", student_name, desc, -amount]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if col_idx == 5:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        current_row += 1
        
    # Payments Total
    ws.cell(row=current_row, column=4, value="Payments Total").font = bold_body_font
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=current_row, column=5, value=-payment_total)
    total_cell.font = bold_body_font
    total_cell.number_format = "$#,##0.00"
    total_cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    total_cell.alignment = Alignment(horizontal="right")
    current_row += 2

    # --- SECTION 6: LEAD CRM PAYMENTS ---
    write_section_header("6. LEAD CRM PAYMENTS")
    write_table_headers(["Date", "Lead ID", "Parent Name", "Status", "Amount"])
    
    from app.models.lead_model import Lead
    leads_in_range = Lead.query.filter(Lead.payment_status == 'Paid', Lead.created_at >= start_date, Lead.created_at <= end_date).all()
    lead_total = 0.0
    
    for lead in leads_in_range:
        parent_name = " / ".join([f"{p.first_name} {p.last_name}" for p in lead.parents]) or "Unknown Parent"
        amount = float(lead.amount or 0)
        lead_total += amount
        
        row_vals = [lead.created_at.strftime('%m/%d/%Y'), f"LEAD#{lead.id}", parent_name, lead.status, -amount]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if col_idx == 5:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        current_row += 1
        
    # Lead Total
    ws.cell(row=current_row, column=4, value="Lead CRM Total").font = bold_body_font
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=current_row, column=5, value=-lead_total)
    total_cell.font = bold_body_font
    total_cell.number_format = "$#,##0.00"
    total_cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    total_cell.alignment = Alignment(horizontal="right")
    current_row += 2

    # --- SECTION 7: LEDGER SUMMARY BALANCE SHEET ---
    write_section_header("7. LEDGER SUMMARY & BALANCE SHEET COMPARISON")
    
    from app.models.financial_model import StudentFinancialAccount
    all_accounts = StudentFinancialAccount.query.all()
    
    debit_sum_start = 0.0
    credit_sum_start = 0.0
    debit_sum_end = 0.0
    credit_sum_end = 0.0
    
    for acc in all_accounts:
        # Sum transactions up to start_date
        inv_start = sum(float(i.total_amount) for i in acc.invoices if i.created_at < start_date)
        ref_start = sum(float(p.amount) for p in acc.payments if p.method == 'Refund' and p.transaction_date < start_date)
        pay_start = sum(float(p.amount) for p in acc.payments if p.method != 'Refund' and p.transaction_date < start_date)
        cred_start = sum(float(c.amount) for c in acc.credits if c.created_at < start_date)
        
        # Query distributions safely
        sub_dist_start = 0.0
        for dist in sub_payment_distributions_by_account(acc.id):
            if dist.transaction.transaction_date < start_date.date():
                sub_dist_start += float(dist.amount)
        
        balance_start = (inv_start + ref_start) - (pay_start + cred_start + sub_dist_start)
        if balance_start > 0:
            debit_sum_start += balance_start
        elif balance_start < 0:
            credit_sum_start += abs(balance_start)
            
        # Sum transactions up to end_date
        inv_end = sum(float(i.total_amount) for i in acc.invoices if i.created_at <= end_date)
        ref_end = sum(float(p.amount) for p in acc.payments if p.method == 'Refund' and p.transaction_date <= end_date)
        pay_end = sum(float(p.amount) for p in acc.payments if p.method != 'Refund' and p.transaction_date <= end_date)
        cred_end = sum(float(c.amount) for c in acc.credits if c.created_at <= end_date)
        
        sub_dist_end = 0.0
        for dist in sub_payment_distributions_by_account(acc.id):
            if dist.transaction.transaction_date <= end_date.date():
                sub_dist_end += float(dist.amount)
        
        balance_end = (inv_end + ref_end) - (pay_end + cred_end + sub_dist_end)
        if balance_end > 0:
            debit_sum_end += balance_end
        elif balance_end < 0:
            credit_sum_end += abs(balance_end)

    summary_labels = [
        ("Ending Sum of Debit Balance (Owed)", debit_sum_end),
        ("Ending Sum of Credit Balance (Prepaid)", credit_sum_end),
        ("Ending Net Balance", debit_sum_end - credit_sum_end),
        ("Change in Debit Balance (Owed Change)", debit_sum_end - debit_sum_start),
        ("Change in Credit Balance (Prepaid Change)", credit_sum_end - credit_sum_start)
    ]
    
    for label, val in summary_labels:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        cell_lbl = ws.cell(row=current_row, column=1, value=label)
        cell_lbl.font = bold_body_font
        cell_lbl.alignment = Alignment(horizontal="right", vertical="center")
        cell_lbl.fill = summary_fill
        cell_lbl.border = thin_border
        
        for c in range(2, 5):
            ws.cell(row=current_row, column=c).fill = summary_fill
            ws.cell(row=current_row, column=c).border = thin_border
            
        cell_val = ws.cell(row=current_row, column=5, value=val)
        cell_val.font = bold_body_font
        cell_val.number_format = "$#,##0.00"
        cell_val.alignment = Alignment(horizontal="right", vertical="center")
        cell_val.fill = summary_fill
        cell_val.border = thin_border
        
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2] or (cell.coordinate in ws.merged_cells):
                continue
            if cell.value:
                val_str = f"${cell.value:,.2f}" if isinstance(cell.value, (int, float)) and cell.number_format == "$#,##0.00" else str(cell.value)
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_path)

def sub_payment_distributions_by_account(account_id):
    from app.models.subsidy_transaction_model import SubsidyPaymentDistribution
    try:
        return SubsidyPaymentDistribution.query.filter_by(student_financial_account_id=account_id).all()
    except Exception:
        return []
